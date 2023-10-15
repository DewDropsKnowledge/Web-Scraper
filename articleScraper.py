import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

def articleScraper(url):
    # URL of the webpage to scrape

    # Send an HTTP GET request to the URL
    response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})

    if response.status_code == 200:
        # Create a BeautifulSoup object to parse the HTML
        soup = BeautifulSoup(response.text, 'html.parser')

        # Find the element containing the article content
        article_element = soup.find("div", class_="post-content")

        # Find the element containing the publication date
        date_element = soup.find("time")

        if article_element and date_element:
            article = article_element.get_text().strip()
            publication_date = date_element.get_text().strip()

            # Load the existing Excel file
            existing_workbook = load_workbook("scraped_articles.xlsx")
            sheet = existing_workbook.active

            # Find the next available row for writing data
            current_row = sheet.max_row + 1

            # Write the data to the Excel file
            sheet[f"A{current_row}"] = publication_date
            sheet[f"B{current_row}"] = article

            # Save the updated Excel file
            existing_workbook.save("scraped_articles.xlsx")

            print("Data has been scraped and appended to 'scraped_articles.xlsx'")
        else:
            print("Failed to extract data from the webpage.")
    else:
        print("Failed to retrieve the webpage. Status code:", response.status_code)

